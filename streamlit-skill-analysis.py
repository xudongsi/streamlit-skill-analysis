import os
import time
from datetime import datetime
from typing import List, Tuple

# å…ˆè®¾ç½®pandasé…ç½®ï¼Œé¿å…ç‰ˆæœ¬å…¼å®¹é—®é¢˜
import pandas as pd

# è®¾ç½®pandasæ˜¾ç¤ºé€‰é¡¹
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
import streamlit as st
from streamlit_autorefresh import st_autorefresh
from streamlit_echarts import st_echarts
import plotly.graph_objects as go
from openpyxl import load_workbook

# å°è¯•å¯¼å…¥psutilï¼Œå¦‚æœå¤±è´¥åˆ™æä¾›æ›¿ä»£æ–¹æ¡ˆ
try:
    import psutil
    PSUTIL_AVAILABLE = True
except ImportError:
    PSUTIL_AVAILABLE = False
    st.sidebar.warning("âš ï¸ psutilåº“æœªå®‰è£…ï¼Œæ–‡ä»¶å ç”¨æ£€æµ‹åŠŸèƒ½ä¸å¯ç”¨")

# -------------------- é¡µé¢é…ç½® --------------------
st.set_page_config(page_title="æŠ€èƒ½è¦†ç›–åˆ†æå¤§å±", layout="wide")

# -------------------- é¡µé¢æ ·å¼ --------------------
PAGE_CSS = """
<style>
body, [data-testid="stAppViewContainer"]{
    background-color:#0d1b2a !important;
    color:#ffffff !important;
}
[data-testid="stSidebar"]{
    background-color:#1b263b !important;
    color:#ffffff !important;
}
div.stButton>button{
    background-color:#4cc9f0 !important;
    color:#000000 !important;
    border-radius:10px;
    height:40px;
    font-weight:700;
    margin:5px 0;
    width:100%;
}
div.stButton>button:hover{
    background-color:#4895ef !important;
    color:#ffffff !important;
}
.metric-card{
    background-color:#1b263b !important;
    padding:20px;
    border-radius:16px;
    text-align:center;
    box-shadow:0 0 15px rgba(0,0,0,0.4);
}
.metric-value{
    font-size:36px;
    font-weight:800;
    color:#4cc9f0 !important;
}
.metric-label{
    font-size:14px;
    color:#cccccc !important;
}
hr{
    border:none;
    border-top:1px solid rgba(255,255,255,.12);
    margin:16px 0;
}
/* çƒ­åŠ›å›¾æ»šåŠ¨å®¹å™¨æ ·å¼ */
.heatmap-container {
    max-height: 700px;
    overflow-y: auto;
    overflow-x: auto;
    border-radius: 8px;
}
/* æ»šåŠ¨æ¡ç¾åŒ– */
.heatmap-container::-webkit-scrollbar {
    width: 8px;
    height: 8px;
}
.heatmap-container::-webkit-scrollbar-thumb {
    background-color: #4cc9f0;
    border-radius: 4px;
}
.heatmap-container::-webkit-scrollbar-track {
    background-color: #1b263b;
}
/* åˆ é™¤æŒ‰é’®æ ·å¼ */
button[data-testid="baseButton-secondary"][key="delete_btn"] {
    background-color: #ff4d4d !important;
    color: white !important;
}
button[data-testid="baseButton-secondary"][key="delete_btn"]:hover {
    background-color: #ff1a1a !important;
}
/* ç¡®è®¤æŒ‰é’®æ ·å¼ */
button[data-testid="baseButton-secondary"][key="confirm_delete"] {
    background-color: #ff6666 !important;
    color: white !important;
}
/* å–æ¶ˆæŒ‰é’®æ ·å¼ */
button[data-testid="baseButton-secondary"][key="cancel_delete"] {
    background-color: #4cc9f0 !important;
    color: black !important;
}
</style>
"""
st.markdown(PAGE_CSS, unsafe_allow_html=True)

SAVE_FILE = "jixiao.xlsx"
# å®‰å…¨æ ¡éªŒï¼šç¡®ä¿æ–‡ä»¶åç¼€æ˜¯xlsxï¼Œé¿å…æ ¼å¼è¯†åˆ«é”™è¯¯
if not SAVE_FILE.endswith(".xlsx"):
    SAVE_FILE += ".xlsx"


# -------------------- å·¥å…·å‡½æ•°ï¼šæ£€æµ‹æ–‡ä»¶æ˜¯å¦è¢«å ç”¨ --------------------
def is_file_locked(file_path):
    if not os.path.exists(file_path):
        return False
    try:
        with open(file_path, 'rb+'):
            return False
    except PermissionError:
        return True
    except Exception:
        return True


# -------------------- æ•°æ®å¯¼å…¥ï¼ˆæ ¸å¿ƒä¿®å¤ï¼šæŒ‡å®šè¯»å–å¼•æ“ï¼‰ --------------------
@st.cache_data(ttl=300)
def load_sheets(file) -> Tuple[List[str], dict]:
    """è¯»å–Excelæ‰€æœ‰å·¥ä½œè¡¨ï¼Œä¿®å¤pandasç‰ˆæœ¬å…¼å®¹é—®é¢˜"""
    if not os.path.exists(file):
        return [], {}

    try:
        # å°è¯•ä½¿ç”¨openpyxlå¼•æ“
        xpd = pd.ExcelFile(file, engine="openpyxl")
    except Exception as e:
        st.sidebar.error(f"âš ï¸ ä½¿ç”¨openpyxlè¯»å–Excelæ–‡ä»¶å¤±è´¥ï¼š{e}")
        # é™çº§å°è¯•è‡ªåŠ¨æ£€æµ‹å¼•æ“
        try:
            xpd = pd.ExcelFile(file)
            st.sidebar.info("ğŸ”§ å·²è‡ªåŠ¨é€‰æ‹©å…¼å®¹å¼•æ“è¯»å–æ–‡ä»¶")
        except Exception as e2:
            st.sidebar.error(f"âŒ æ‰€æœ‰å¼•æ“è¯»å–å¤±è´¥ï¼š{e2}")
            return [], {}

    frames = {}
    for s in xpd.sheet_names:
        try:
            # å°è¯•ä½¿ç”¨openpyxlè¯»å–ï¼Œå¦‚æœå¤±è´¥åˆ™ä½¿ç”¨é»˜è®¤
            try:
                df0 = pd.read_excel(xpd, sheet_name=s, engine="openpyxl")
            except:
                df0 = pd.read_excel(xpd, sheet_name=s)
                
            if df0.empty:
                continue
                
            # æ£€æŸ¥å¿…è¦åˆ—æ˜¯å¦å­˜åœ¨
            required_cols = ["æ˜ç»†", "å‘˜å·¥", "å€¼"]
            missing_cols = [col for col in required_cols if col not in df0.columns]
            if missing_cols:
                st.sidebar.warning(f"âš ï¸ è¡¨ {s} ç¼ºå°‘åˆ— {missing_cols}ï¼Œå·²è·³è¿‡ã€‚")
                continue

            # è§£æåˆ†ç»„è¡Œ
            if not df0.empty and df0.iloc[0, 0] == "åˆ†ç»„":
                groups = df0.iloc[0, 1:].tolist()
                df0 = df0.drop(0).reset_index(drop=True)
                emp_cols = [c for c in df0.columns if c not in ["æ˜ç»†", "æ•°é‡æ€»å’Œ", "ç¼–å·"]]
                group_map = {emp: groups[i] if i < len(groups) else None for i, emp in enumerate(emp_cols)}
                df_long = df0.melt(
                    id_vars=["æ˜ç»†", "æ•°é‡æ€»å’Œ"] if "æ•°é‡æ€»å’Œ" in df0.columns else ["æ˜ç»†"],
                    value_vars=emp_cols,
                    var_name="å‘˜å·¥",
                    value_name="å€¼"
                )
                df_long["åˆ†ç»„"] = df_long["å‘˜å·¥"].map(group_map)
                frames[s] = df_long
            else:
                frames[s] = df0
        except Exception as e:
            st.sidebar.error(f"âŒ è¯»å– {s} æ—¶å‡ºé”™: {e}")
    return xpd.sheet_names, frames


# -------------------- ä¼˜åŒ–åçš„åˆ é™¤å·¥ä½œè¡¨å‡½æ•° --------------------
def delete_sheet_optimized(file_path, sheet_name):
    if not os.path.exists(file_path):
        return False, "âŒ æ–‡ä»¶ä¸å­˜åœ¨"

    if is_file_locked(file_path):
        return False, "âŒ æ–‡ä»¶è¢«å ç”¨ï¼ˆå¯èƒ½Excelå·²æ‰“å¼€ï¼‰ï¼Œè¯·å…³é—­Excelåé‡è¯•"

    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return False, "âŒ å·¥ä½œè¡¨ä¸å­˜åœ¨"

        wb.remove(wb[sheet_name])
        wb.save(file_path)
        wb.close()

        return True, f"âœ… æˆåŠŸåˆ é™¤å·¥ä½œè¡¨: {sheet_name}"
    except PermissionError:
        return False, "âŒ æƒé™ä¸è¶³ï¼Œæ— æ³•åˆ é™¤å·¥ä½œè¡¨ï¼ˆè¯·æ£€æŸ¥æ–‡ä»¶æ˜¯å¦åªè¯»ï¼‰"
    except Exception as e:
        return False, f"âŒ åˆ é™¤å¤±è´¥: {str(e)}"


# -------------------- æ–‡ä»¶è¯»å– --------------------
sheets, sheet_frames = load_sheets(SAVE_FILE)

# åˆå§‹åŒ–ï¼šæ–‡ä»¶ä¸å­˜åœ¨æ—¶åˆ›å»ºç©ºæ–‡ä»¶
if not os.path.exists(SAVE_FILE):
    try:
        with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
            pd.DataFrame(columns=["æ˜ç»†", "æ•°é‡æ€»å’Œ", "å‘˜å·¥", "å€¼", "åˆ†ç»„"]).to_excel(
                writer, sheet_name="ç¤ºä¾‹_2025_01", index=False
            )
        sheets, sheet_frames = load_sheets(SAVE_FILE)
        st.sidebar.success(f"âœ… å·²åˆ›å»ºåˆå§‹æ–‡ä»¶ {SAVE_FILE}")
    except Exception as e:
        st.sidebar.error(f"âŒ åˆ›å»ºåˆå§‹æ–‡ä»¶å¤±è´¥ï¼š{e}")
elif not sheets:
    st.sidebar.warning("âš ï¸ æ–‡ä»¶å­˜åœ¨ä½†æ— æœ‰æ•ˆå·¥ä½œè¡¨ï¼Œå·²åˆ›å»ºç¤ºä¾‹æ•°æ®")
    sheet_frames = {
        "ç¤ºä¾‹_2025_01": pd.DataFrame({
            "æ˜ç»†": ["ä»»åŠ¡A", "ä»»åŠ¡B", "ä»»åŠ¡C"],
            "æ•°é‡æ€»å’Œ": [3, 2, 5],
            "å‘˜å·¥": ["å¼ ä¸‰", "æå››", "ç‹äº”"],
            "å€¼": [1, 1, 1],
            "åˆ†ç»„": ["A8", "B7", "VN"]
        })
    }
    sheets = ["ç¤ºä¾‹_2025_01"]
else:
    st.sidebar.success(f"âœ… å·²åŠ è½½åº“æ–‡ä»¶ {SAVE_FILE}ï¼ˆå…±{len(sheets)}ä¸ªå·¥ä½œè¡¨ï¼‰")

# ---------- ğŸ§  è‡ªåŠ¨æ£€æµ‹å¹¶ä¿®å¤æ•°é‡æ€»å’Œ ----------
repaired_count = 0
repaired_frames = {}
for sheet_name, df0 in sheet_frames.items():
    if "æ˜ç»†" in df0.columns and "å€¼" in df0.columns:
        if "æ•°é‡æ€»å’Œ" not in df0.columns or df0["æ•°é‡æ€»å’Œ"].isnull().any():
            repaired = True
        else:
            try:
                true_sum = df0.groupby("æ˜ç»†")["å€¼"].sum().reset_index()
                merged = df0.merge(true_sum, on="æ˜ç»†", how="left", suffixes=("", "_çœŸå®"))
                repaired = not merged["æ•°é‡æ€»å’Œ"].equals(merged["å€¼_çœŸå®"])
            except:
                repaired = True

        if repaired:
            repaired_count += 1
            sum_df = (
                df0.groupby("æ˜ç»†", as_index=False)["å€¼"].sum()
                .rename(columns={"å€¼": "æ•°é‡æ€»å’Œ"})
            )
            df0 = df0.drop(columns=["æ•°é‡æ€»å’Œ"], errors="ignore")
            df0 = df0.merge(sum_df, on="æ˜ç»†", how="left")
            repaired_frames[sheet_name] = df0

if repaired_frames:
    try:
        with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
            for sn in sheets:
                if sn in repaired_frames:
                    repaired_df = repaired_frames[sn]
                    repaired_df.to_excel(writer, sheet_name=sn, index=False)
                    sheet_frames[sn] = repaired_df
                else:
                    try:
                        df_original = pd.read_excel(SAVE_FILE, sheet_name=sn, engine="openpyxl")
                        df_original.to_excel(writer, sheet_name=sn, index=False)
                    except:
                        df_original = pd.read_excel(SAVE_FILE, sheet_name=sn)
                        df_original.to_excel(writer, sheet_name=sn, index=False)
        st.cache_data.clear()
        st.sidebar.info(f"ğŸ”§ å·²è‡ªåŠ¨ä¿®å¤ {repaired_count} å¼ è¡¨çš„æ•°é‡æ€»å’Œåˆ—")
    except Exception as e:
        st.sidebar.error(f"âŒ ä¿®å¤æ•°é‡æ€»å’Œå¤±è´¥ï¼š{e}")

# -------------------- æ™ºèƒ½åŒ–æ–°å¢æœˆä»½/å­£åº¦ --------------------
st.sidebar.markdown("### ğŸ“… æ–°å¢æ•°æ®æ—¶é—´ç‚¹")
current_year = datetime.now().year
year = st.sidebar.selectbox("é€‰æ‹©å¹´ä»½", list(range(current_year - 2, current_year + 2)), index=2)
mode = st.sidebar.radio("æ—¶é—´ç±»å‹", ["æœˆä»½", "å­£åº¦"], horizontal=True)

if mode == "æœˆä»½":
    month = st.sidebar.selectbox("é€‰æ‹©æœˆä»½", list(range(1, 13)))
    new_sheet_name = f"{year}_{month:02d}"
else:
    quarter = st.sidebar.selectbox("é€‰æ‹©å­£åº¦", ["Q1", "Q2", "Q3", "Q4"])
    new_sheet_name = f"{year}_{quarter}"

if st.sidebar.button("åˆ›å»ºæ–°çš„æ—¶é—´ç‚¹"):
    if new_sheet_name in sheets:
        st.sidebar.error(f"âŒ æ—¶é—´ç‚¹ {new_sheet_name} å·²å­˜åœ¨ï¼")
    else:
        try:
            base_df = pd.DataFrame(columns=["æ˜ç»†", "æ•°é‡æ€»å’Œ", "å‘˜å·¥", "å€¼", "åˆ†ç»„"])
            prev_sheets = sorted([s for s in sheets if "_" in s and s < new_sheet_name])
            if prev_sheets:
                prev_name = prev_sheets[-1]
                base_df = sheet_frames.get(prev_name, base_df).copy()
                st.sidebar.info(f"ğŸ”§ å·²ä»æœ€è¿‘æ—¶é—´ç‚¹ {prev_name} è‡ªåŠ¨ç»§æ‰¿æ•°æ®")
            else:
                st.sidebar.info("ğŸ”§ æœªæ‰¾åˆ°ä¸ŠæœŸæ•°æ®ï¼Œåˆ›å»ºç©ºç™½æ¨¡æ¿")

            # å†™å…¥æ—¶æŒ‡å®šå¼•æ“
            with pd.ExcelWriter(SAVE_FILE, mode="a", engine="openpyxl") as writer:
                base_df.to_excel(writer, sheet_name=new_sheet_name, index=False)

            st.cache_data.clear()
            sheets, sheet_frames = load_sheets(SAVE_FILE)
            st.sidebar.success(f"âœ… å·²åˆ›å»ºæ–°æ—¶é—´ç‚¹: {new_sheet_name}")

        except Exception as e:
            st.sidebar.error(f"âŒ åˆ›å»ºå¤±è´¥ï¼š{e}")

# -------------------- ä¼˜åŒ–åçš„åˆ é™¤å·¥ä½œè¡¨åŠŸèƒ½ --------------------
st.sidebar.markdown("### ğŸ—‘ï¸ åˆ é™¤æ—¶é—´ç‚¹")
if sheets:
    sheet_to_delete = st.sidebar.selectbox("é€‰æ‹©è¦åˆ é™¤çš„æ—¶é—´ç‚¹", sheets, key="delete_sheet_select")

    if len(sheets) == 1:
        st.sidebar.warning("âš ï¸ è‡³å°‘ä¿ç•™ä¸€ä¸ªå·¥ä½œè¡¨ï¼Œæ— æ³•åˆ é™¤")
    else:
        if "delete_confirm" not in st.session_state:
            st.session_state.delete_confirm = False

        if not st.session_state.delete_confirm:
            if st.sidebar.button("åˆ é™¤é€‰ä¸­æ—¶é—´ç‚¹", key="delete_btn", help="åˆ é™¤åä¸å¯æ¢å¤"):
                st.session_state.delete_confirm = True
        else:
            st.sidebar.warning(f"âš ï¸ ç¡®è®¤åˆ é™¤ã€{sheet_to_delete}ã€‘ï¼Ÿæ­¤æ“ä½œä¸å¯æ¢å¤ï¼")
            col1, col2 = st.sidebar.columns(2)
            with col1:
                if st.button("ç¡®è®¤åˆ é™¤", key="confirm_delete"):
                    success, msg = delete_sheet_optimized(SAVE_FILE, sheet_to_delete)
                    st.sidebar.warning(msg)
                    if success:
                        st.cache_data.clear()
                        sheets, sheet_frames = load_sheets(SAVE_FILE)
                        st.session_state.delete_confirm = False
                        st.rerun()
            with col2:
                if st.button("å–æ¶ˆ", key="cancel_delete"):
                    st.session_state.delete_confirm = False

# -------------------- ğŸ§® ä¸€é”®æ›´æ–°æ‰€æœ‰æ•°é‡æ€»å’Œ --------------------
st.sidebar.markdown("### ğŸ”§ æ•°æ®ä¿®å¤å·¥å…·")

if st.sidebar.button("ğŸ§® ä¸€é”®æ›´æ–°æ‰€æœ‰æ•°é‡æ€»å’Œ"):
    try:
        if not os.path.exists(SAVE_FILE):
            st.sidebar.warning("æœªæ‰¾åˆ°æ–‡ä»¶ jixiao.xlsx")
        else:
            try:
                xls = pd.ExcelFile(SAVE_FILE, engine="openpyxl")
            except:
                xls = pd.ExcelFile(SAVE_FILE)
                
            updated_frames = {}
            for sheet_name in xls.sheet_names:
                try:
                    df0 = pd.read_excel(xls, sheet_name=sheet_name, engine="openpyxl")
                except:
                    df0 = pd.read_excel(xls, sheet_name=sheet_name)
                    
                if "æ˜ç»†" in df0.columns and "å€¼" in df0.columns:
                    sum_df = (
                        df0.groupby("æ˜ç»†", as_index=False)["å€¼"].sum()
                        .rename(columns={"å€¼": "æ•°é‡æ€»å’Œ"})
                    )
                    df0 = df0.drop(columns=["æ•°é‡æ€»å’Œ"], errors="ignore")
                    df0 = df0.merge(sum_df, on="æ˜ç»†", how="left")
                    updated_frames[sheet_name] = df0

            with pd.ExcelWriter(SAVE_FILE, engine="openpyxl") as writer:
                for sheet_name, df0 in updated_frames.items():
                    df0.to_excel(writer, sheet_name=sheet_name, index=False)

            st.cache_data.clear()
            sheets, sheet_frames = load_sheets(SAVE_FILE)
            st.sidebar.success("âœ… æ‰€æœ‰å·¥ä½œè¡¨çš„æ•°é‡æ€»å’Œå·²é‡æ–°è®¡ç®—å¹¶æ›´æ–°ï¼")

    except Exception as e:
        st.sidebar.error(f"âŒ æ›´æ–°å¤±è´¥ï¼š{e}")

# -------------------- æ—¶é—´ç‚¹é€‰æ‹©ä¼˜åŒ– --------------------
st.sidebar.markdown("### ğŸ“‹ æ•°æ®ç­›é€‰")
years_available = sorted(list({s.split("_")[0] for s in sheets if "_" in s}))
year_choice = st.sidebar.selectbox("ç­›é€‰å¹´ä»½", ["å…¨éƒ¨å¹´ä»½"] + years_available)

if year_choice == "å…¨éƒ¨å¹´ä»½":
    time_candidates = sorted(sheets)
else:
    time_candidates = sorted([s for s in sheets if s.startswith(year_choice)])

if not time_candidates:
    st.warning(f"âš ï¸ æš‚æ— ç¬¦åˆæ¡ä»¶çš„æ•°æ®ï¼Œè¯·å…ˆåˆ›å»ºæœˆä»½æˆ–å­£åº¦ã€‚")
    time_choice = []
else:
    default_choice = time_candidates[:2] if len(time_candidates) >= 2 else time_candidates[:1]
    time_choice = st.sidebar.multiselect("é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæ”¯æŒè·¨å¹´ä»½å¯¹æ¯”ï¼‰",
                                         time_candidates,
                                         default=default_choice)

# -------------------- åˆ†ç»„é€‰æ‹© --------------------
all_groups = pd.concat(sheet_frames.values())["åˆ†ç»„"].dropna().unique().tolist() if sheet_frames else []
selected_groups = st.sidebar.multiselect("é€‰æ‹©åˆ†ç»„", all_groups, default=all_groups)

# -------------------- è§†å›¾é€‰æ‹© --------------------
sections_names = [
    "äººå‘˜å®Œæˆä»»åŠ¡æ•°é‡æ’å",
    "ä»»åŠ¡å¯¹æ¯”ï¼ˆå †å æŸ±çŠ¶å›¾ï¼‰",
    "ä»»åŠ¡-äººå‘˜çƒ­åŠ›å›¾"
]
view = st.sidebar.radio("åˆ‡æ¢è§†å›¾", ["ç¼–è¾‘æ•°æ®", "å¤§å±è½®æ’­", "å•é¡µæ¨¡å¼", "æ˜¾ç¤ºæ‰€æœ‰è§†å›¾", "èƒ½åŠ›åˆ†æ"])


# -------------------- æ•°æ®åˆå¹¶ --------------------
def get_merged_df(keys: List[str], groups: List[str]) -> pd.DataFrame:
    dfs = []
    for k in keys:
        df0 = sheet_frames.get(k)
        if df0 is not None:
            if groups and "åˆ†ç»„" in df0.columns:
                df0 = df0[df0["åˆ†ç»„"].isin(groups)]
            dfs.append(df0)
    if not dfs:
        st.warning("âš ï¸ å½“å‰é€‰æ‹©æ²¡æœ‰æ•°æ®ï¼Œè¯·æ£€æŸ¥æ—¶é—´ç‚¹æˆ–åˆ†ç»„é€‰æ‹©ã€‚")
        return pd.DataFrame()
    return pd.concat(dfs, axis=0, ignore_index=True)


df = get_merged_df(time_choice, selected_groups)


# -------------------- å›¾è¡¨å‡½æ•° --------------------
def chart_total(df0):
    if df0.empty:
        return go.Figure()
    
    df0 = df0[df0["æ˜ç»†"] != "åˆ†æ•°æ€»å’Œ"]
    if df0.empty:
        return go.Figure()
        
    emp_stats = df0.groupby("å‘˜å·¥")["å€¼"].sum().sort_values(ascending=False).reset_index()
    fig = go.Figure(go.Bar(
        x=emp_stats["å‘˜å·¥"],
        y=emp_stats["å€¼"],
        text=emp_stats["å€¼"],
        textposition="outside",
        hovertemplate="å‘˜å·¥: %{x}<br>å®Œæˆæ€»å€¼: %{y}<extra></extra>"
    ))
    fig.update_layout(template="plotly_dark", xaxis_title="å‘˜å·¥", yaxis_title="å®Œæˆæ€»å€¼")
    return fig


def chart_stack(df0):
    if df0.empty:
        return go.Figure()
    
    df0 = df0[df0["æ˜ç»†"] != "åˆ†æ•°æ€»å’Œ"]
    if df0.empty:
        return go.Figure()
        
    df_pivot = df0.pivot_table(index="æ˜ç»†", columns="å‘˜å·¥", values="å€¼", aggfunc="sum", fill_value=0)
    fig = go.Figure()
    for emp in df_pivot.columns:
        fig.add_trace(go.Bar(x=df_pivot.index, y=df_pivot[emp], name=emp))
    fig.update_layout(barmode="stack", template="plotly_dark", xaxis_title="ä»»åŠ¡", yaxis_title="å®Œæˆå€¼")
    return fig


def chart_heat(df0):
    if df0.empty:
        return {}
    
    df0 = df0[df0["æ˜ç»†"] != "åˆ†æ•°æ€»å’Œ"]
    if df0.empty:
        return {}
        
    tasks = df0["æ˜ç»†"].unique().tolist()
    emps = df0["å‘˜å·¥"].unique().tolist()
    data = []
    for i, t in enumerate(tasks):
        for j, e in enumerate(emps):
            v = int(df0[(df0["æ˜ç»†"] == t) & (df0["å‘˜å·¥"] == e)]["å€¼"].sum())
            data.append([j, i, v])
    return {
        "backgroundColor": "transparent",
        "tooltip": {"position": "top"},
        "xAxis": {"type": "category", "data": emps, "axisLabel": {"color": "#fff", "rotate": 45}},
        "yAxis": {"type": "category", "data": tasks, "axisLabel": {"color": "#fff"}},
        "visualMap": {"min": 0, "max": max([d[2] for d in data]) if data else 1, "show": True,
                      "inRange": {"color": ["#ff4d4d", "#4caf50"]}, "textStyle": {"color": "#fff"}},
        "series": [{"type": "heatmap", "data": data, "emphasis": {"itemStyle": {"shadowBlur": 10}}}]
    }


# -------------------- å¡ç‰‡æ˜¾ç¤º --------------------
def show_cards(df0):
    if df0.empty:
        st.info("ğŸ“Š å½“å‰æ— æ•°æ®ï¼Œè¯·é€‰æ‹©æ—¶é—´ç‚¹å’Œåˆ†ç»„")
        return
        
    df0 = df0[df0["æ˜ç»†"] != "åˆ†æ•°æ€»å’Œ"]
    if df0.empty:
        st.info("ğŸ“Š å½“å‰æ— æ•°æ®ï¼ˆä¸åŒ…å«åˆ†æ•°æ€»å’Œï¼‰")
        return

    total_tasks = df0["æ˜ç»†"].nunique()
    total_people = df0["å‘˜å·¥"].nunique()
    ps = df0.groupby("å‘˜å·¥")["å€¼"].sum()
    top_person = ps.idxmax() if not ps.empty else "æ— "
    avg_score = round(ps.mean(), 1) if not ps.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(
        f"<div class='metric-card'><div class='metric-value'>{total_tasks}</div><div class='metric-label'>ä»»åŠ¡æ•°</div></div>",
        unsafe_allow_html=True)
    c2.markdown(
        f"<div class='metric-card'><div class='metric-value'>{total_people}</div><div class='metric-label'>äººæ•°</div></div>",
        unsafe_allow_html=True)
    c3.markdown(
        f"<div class='metric-card'><div class='metric-value'>{top_person}</div><div class='metric-label'>è¦†ç›–ç‡æœ€é«˜</div></div>",
        unsafe_allow_html=True)
    c4.markdown(
        f"<div class='metric-card'><div class='metric-value'>{avg_score}</div><div class='metric-label'>å¹³å‡æ•°</div></div>",
        unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)


# -------------------- å®šä¹‰é²œè‰³çš„é¢œè‰²åˆ—è¡¨ --------------------
BRIGHT_COLORS = [
    "#FF0000", "#00FF00", "#0000FF", "#FFA500", "#800080",
    "#00FFFF", "#FFC0CB", "#FFFF00", "#008080", "#FF00FF"
]

# -------------------- ä¸»é¡µé¢ --------------------
st.title("ğŸ“Š æŠ€èƒ½è¦†ç›–åˆ†æå¤§å±")

if view == "ç¼–è¾‘æ•°æ®":
    if not time_choice:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæœˆæˆ–å­£ï¼‰åå†ç¼–è¾‘æ•°æ®")
    elif len(time_choice) > 1:
        st.warning("âš ï¸ ç¼–è¾‘æ•°æ®æ—¶ä»…æ”¯æŒé€‰æ‹©å•ä¸ªæ—¶é—´ç‚¹ï¼Œè¯·é‡æ–°é€‰æ‹©ï¼")
    else:
        show_cards(df)
        st.info("ä½ å¯ä»¥ç›´æ¥ç¼–è¾‘ä¸‹é¢çš„è¡¨æ ¼ï¼Œä¿®æ”¹å®Œæˆåç‚¹å‡»ã€ä¿å­˜ã€‘æŒ‰é’®ã€‚")

        sheet_name = time_choice[0]
        try:
            try:
                original_df = pd.read_excel(SAVE_FILE, sheet_name=sheet_name, engine="openpyxl")
            except:
                original_df = pd.read_excel(SAVE_FILE, sheet_name=sheet_name)
                
            edited_df = st.data_editor(df, num_rows="dynamic", use_container_width=True)

            if st.button("ğŸ’¾ ä¿å­˜ä¿®æ”¹åˆ°åº“é‡Œ"):
                try:
                    if selected_groups and "åˆ†ç»„" in original_df.columns:
                        mask = original_df["åˆ†ç»„"].isin(selected_groups)
                        original_df = original_df[~mask].reset_index(drop=True)
                        final_df = pd.concat([original_df, edited_df], ignore_index=True)
                    else:
                        final_df = edited_df.copy()

                    if "æ˜ç»†" in final_df.columns and "å€¼" in final_df.columns:
                        sum_df = (
                            final_df.groupby("æ˜ç»†", as_index=False)["å€¼"].sum()
                            .rename(columns={"å€¼": "æ•°é‡æ€»å’Œ"})
                        )
                        final_df = final_df.drop(columns=["æ•°é‡æ€»å’Œ"], errors="ignore")
                        final_df = final_df.merge(sum_df, on="æ˜ç»†", how="left")

                    with pd.ExcelWriter(SAVE_FILE, mode="a", if_sheet_exists="replace", engine="openpyxl") as writer:
                        final_df.to_excel(writer, sheet_name=sheet_name, index=False)

                    st.cache_data.clear()
                    sheets, sheet_frames = load_sheets(SAVE_FILE)
                    st.success(f"âœ… ä¿®æ”¹å·²ä¿å­˜åˆ° {SAVE_FILE} ({sheet_name})ï¼Œä»…æ›´æ–°é€‰ä¸­åˆ†ç»„æ•°æ®")
                except Exception as e:
                    st.error(f"ä¿å­˜å¤±è´¥ï¼š{e}")
        except Exception as e:
            st.error(f"âŒ åŠ è½½ç¼–è¾‘æ•°æ®å¤±è´¥ï¼š{e}")

elif view == "å¤§å±è½®æ’­":
    if not time_choice:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæœˆæˆ–å­£ï¼‰åæŸ¥çœ‹å¤§å±è½®æ’­")
    else:
        st_autorefresh(interval=10000, key="aut")
        show_cards(df)
        secs = [("å®Œæˆæ’å", chart_total(df)),
                ("ä»»åŠ¡å¯¹æ¯”", chart_stack(df)),
                ("çƒ­åŠ›å›¾", chart_heat(df))]
        current_index = int(time.time() / 10) % len(secs)
        t, op = secs[current_index]
        st.subheader(t)
        if isinstance(op, go.Figure):
            st.plotly_chart(op, use_container_width=True)
        else:
            st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
            st_echarts(op, height=f"{max(600, len(df['æ˜ç»†'].unique()) * 25)}px", theme="dark")
            st.markdown('</div>', unsafe_allow_html=True)

elif view == "å•é¡µæ¨¡å¼":
    if not time_choice:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæœˆæˆ–å­£ï¼‰åæŸ¥çœ‹å•é¡µæ¨¡å¼")
    else:
        show_cards(df)
        choice = st.sidebar.selectbox("å•é¡µæŸ¥çœ‹", sections_names, index=0)
        mapping = {
            "äººå‘˜å®Œæˆä»»åŠ¡æ•°é‡æ’å": chart_total(df),
            "ä»»åŠ¡å¯¹æ¯”ï¼ˆå †å æŸ±çŠ¶å›¾ï¼‰": chart_stack(df),
            "ä»»åŠ¡-äººå‘˜çƒ­åŠ›å›¾": chart_heat(df)
        }
        chart_func = mapping.get(choice, chart_total(df))
        if isinstance(chart_func, go.Figure):
            st.plotly_chart(chart_func, use_container_width=True)
        else:
            st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
            st_echarts(chart_func, height=f"{max(600, len(df['æ˜ç»†'].unique()) * 25)}px", theme="dark")
            st.markdown('</div>', unsafe_allow_html=True)

elif view == "æ˜¾ç¤ºæ‰€æœ‰è§†å›¾":
    if not time_choice:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæœˆæˆ–å­£ï¼‰åæŸ¥çœ‹æ‰€æœ‰è§†å›¾")
    else:
        show_cards(df)
        charts = [("å®Œæˆæ’å", chart_total(df)),
                  ("ä»»åŠ¡å¯¹æ¯”ï¼ˆå †å æŸ±çŠ¶å›¾ï¼‰", chart_stack(df)),
                  ("çƒ­å›¾", chart_heat(df))]
        for label, f in charts:
            st.subheader(label)
            if isinstance(f, go.Figure):
                st.plotly_chart(f, use_container_width=True)
            else:
                st.markdown('<div class="heatmap-container">', unsafe_allow_html=True)
                st_echarts(f, height=f"{max(600, len(df['æ˜ç»†'].unique()) * 25)}px", theme="dark")
                st.markdown('</div>', unsafe_allow_html=True)

elif view == "èƒ½åŠ›åˆ†æ":
    if not time_choice:
        st.warning("âš ï¸ è¯·åœ¨å·¦ä¾§é€‰æ‹©æ—¶é—´ç‚¹ï¼ˆæœˆæˆ–å­£ï¼‰åæŸ¥çœ‹èƒ½åŠ›åˆ†æ")
    else:
        st.subheader("ğŸ“ˆ èƒ½åŠ›åˆ†æ")
        employees = df["å‘˜å·¥"].unique().tolist() if not df.empty else []
        selected_emps = st.sidebar.multiselect("é€‰æ‹©å‘˜å·¥ï¼ˆå›¾1æ˜¾ç¤ºï¼‰", employees, default=employees)
        tasks = df["æ˜ç»†"].unique().tolist() if not df.empty else []

        fig1, fig2, fig3 = go.Figure(), go.Figure(), go.Figure()
        sheet_color_map = {}
        for idx, sheet in enumerate(time_choice):
            sheet_color_map[sheet] = BRIGHT_COLORS[idx % len(BRIGHT_COLORS)]

        emp_color_idx = 0
        for sheet in time_choice:
            df_sheet = get_merged_df([sheet], selected_groups)
            if df_sheet.empty:
                continue
                
            df_sheet = df_sheet[df_sheet["æ˜ç»†"] != "åˆ†æ•°æ€»å’Œ"]
            if df_sheet.empty:
                continue
                
            try:
                df_pivot = df_sheet.pivot(index="æ˜ç»†", columns="å‘˜å·¥", values="å€¼").fillna(0)
            except:
                continue

            for emp in selected_emps:
                if emp in df_pivot.columns:
                    fig1.add_trace(go.Scatter(
                        x=tasks,
                        y=df_pivot[emp].reindex(tasks, fill_value=0),
                        mode="lines+markers",
                        name=f"{sheet}-{emp}",
                        line=dict(color=BRIGHT_COLORS[emp_color_idx % len(BRIGHT_COLORS)], width=3),
                        marker=dict(size=8)
                    ))
                    emp_color_idx += 1

            fig2.add_trace(go.Scatter(
                x=tasks,
                y=df_pivot.sum(axis=1).reindex(tasks, fill_value=0),
                mode="lines+markers",
                name=sheet,
                line=dict(color=sheet_color_map[sheet], width=3),
                marker=dict(size=8)
            ))

            fig3.add_trace(go.Bar(
                x=df_pivot.columns.tolist(),
                y=df_pivot.sum(axis=0).tolist(),
                name=sheet,
                marker=dict(color=sheet_color_map[sheet]),
                width=0.3,
            ))

        if len(fig1.data) > 0:
            fig1.update_layout(
                title="å‘˜å·¥ä»»åŠ¡å®Œæˆæƒ…å†µ",
                template="plotly_dark",
                font=dict(size=12),
                legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
                height=500
            )
            st.plotly_chart(fig1, use_container_width=True)
        else:
            st.info("ğŸ“Š æš‚æ— å‘˜å·¥å®Œæˆæƒ…å†µæ•°æ®")

        if len(fig2.data) > 0:
            fig2.update_layout(
                title="ä»»åŠ¡æ•´ä½“å®Œæˆåº¦è¶‹åŠ¿",
                template="plotly_dark",
                font=dict(size=12),
                legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
                height=500
            )
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("ğŸ“ˆ æš‚æ— ä»»åŠ¡å®Œæˆè¶‹åŠ¿æ•°æ®")

        if len(fig3.data) > 0:
            fig3.update_layout(
                title="å‘˜å·¥æ•´ä½“å®Œæˆåº¦å¯¹æ¯”",
                template="plotly_dark",
                font=dict(size=12),
                barmode="group",
                bargap=0.25,
                bargroupgap=0.005,
                legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5),
                height=600,
                xaxis=dict(
                    tickangle=45,
                    tickfont=dict(size=10)
                ),
                yaxis=dict(
                    tickfont=dict(size=10)
                )
            )
            st.plotly_chart(fig3, use_container_width=True)
        else:
            st.info("ğŸ‘¥ æš‚æ— å‘˜å·¥å®Œæˆåº¦å¯¹æ¯”æ•°æ®")
